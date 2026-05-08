from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

import config


def write_excel(records, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "商品数据"

    columns = config.OUTPUT_COLUMNS
    _write_header(ws, columns)
    _write_data(ws, records, columns)
    _auto_width(ws, columns)

    wb.save(output_path)


def _write_header(ws, columns):
    header_font = Font(bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36


def _write_data(ws, records, columns):
    data_align = Alignment(horizontal="center", vertical="center")

    for row_idx, record in enumerate(records, 2):
        row_data = record.to_row(columns)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align


def _auto_width(ws, columns):
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name.replace("\n", ""))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            for cell_val in row:
                if cell_val:
                    cell_str = str(cell_val)
                    char_len = 0
                    for ch in cell_str:
                        char_len += 2 if '一' <= ch <= '鿿' else 1
                    max_len = max(max_len, char_len)

        width = min(max_len + 4, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
